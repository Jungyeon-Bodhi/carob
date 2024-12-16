#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Sep 17 15:04:49 2024

@author: Bodhi Global Analysis (Jungyeon Lee)
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
import seaborn as sns
from PIL import Image
from pandas.plotting import table
from IPython.display import clear_output
import warnings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tools.tools import add_constant
from scipy.stats import normaltest
from statsmodels.stats.diagnostic import lilliefors
import statsmodels.api as sm
from scipy import stats
from scipy.stats import f_oneway
from statsmodels.formula.api import ols

warnings.filterwarnings("ignore")
plt.rcParams['figure.dpi'] = 600

bodhi_blue = (0.0745, 0.220, 0.396)
bodhi_grey = (0.247, 0.29, 0.322)
bodhi_primary_1 = (0.239, 0.38, 0.553)
bodhi_secondary = (0.133, 0.098, 0.42)
bodhi_tertiary = (0.047, 0.396, 0.298)
bodhi_complement = (0.604, 0.396, 0.071)

class Data_analysis:

    def __init__(self, name, indicators):
        """
        - Initialise the data analysis class

        name: str, Name of the project
        indicators: list, List of indicators
        """
        self.name = name
        self.indicators = indicators

    def count(self, df, var, index_name):
        """
        - To generate a table showing the count and percentage of the indicator
        df: Dataframe, Dataframe of this project
        var: list, Variables related to the indicator
        index_name: str, Index name for the new count dataframe
        """
        count = df[var].value_counts()
        count_df = pd.DataFrame({'Count': count})
        count_df['Percentage'] = round(count_df['Count'] / count_df['Count'].sum() * 100, 1)
        count_df.index.name = index_name
        return count_df

    def multi_table(self, df, columns, categories, column_labels, index_name, change = None):
        """
        - To generate a multi-table showing the count and percentage of the indicator (For KAP, will be updated soon)
        df: Dataframe, Dataframe of this project
        columns: list, Variables related to the indicator
        categories: list, Categories of the indices
        columns_labels: list, Labels of the columns
        index_name: str, Name of the dataframe
        change: list, New indices
        """
        table = pd.DataFrame(index=categories)
        for col in columns:
            table[col] = df[col].value_counts().reindex(categories, fill_value=0)
        if column_labels is not None:
            table.columns = column_labels
        if change is not None:
            table.index = change[:len(table)]
        column_sums = table.sum(axis=0)
        percentage_table = table.div(column_sums, axis=1) * 100

        for i, idx in enumerate(table.index):
            table.loc[f'{idx}(%)'] = percentage_table.loc[idx]
        return table

    def chi2_test(self, file_path):
        """
        - To run chi-2 test and save the result
        file_path: str, Directory where the chi2 test results will be saved
        """            
        def chi2(df2, dependent_var, col, alpha=0.05):
            try:
                df2_c = df2[dependent_var].dropna()
                df2_ca = df2[col].dropna()
                df2_clean = pd.concat([df2_c, df2_ca], axis=1).dropna()

                df2_clean[dependent_var] = df2_clean[dependent_var].squeeze()
                df2_clean[col] = df2_clean[col].squeeze()
                
                if len(df2_clean[dependent_var]) != len(df2_clean) or len(df2_clean[col]) != len(df2_clean):
                    raise ValueError("Length of columns does not match DataFrame length.")

                df2_clean[col] = df2_clean[col].astype('category')
                df2_clean[dependent_var] = df2_clean[dependent_var].astype('category')
                contingency_table = pd.crosstab(df2_clean[dependent_var], df2_clean[col])
                print(len(contingency_table))
                if contingency_table.empty:
                    raise ValueError("Contingency table is empty")
                chi2, p_value, _, _ = stats.chi2_contingency(contingency_table)
        
                print(f"Chi-square test statistic: {chi2}")
                print(f"P-value: {p_value}")
                var = f'Chi-square Test - {col}'
        
                if p_value < alpha:
                    print(f"Variable: {col} | There is a significant association between {dependent_var} and {col}")
                    print("")
                else:
                    print(f"Variable: {col} | There is not a significant association between {dependent_var} and {col}")
                    print("")
            
                result_df = pd.DataFrame({'Test': [var],'Statistics': [chi2],'P-value': [p_value]})
                return result_df
            except ValueError as e:
                print(f"Error: {e}")
        
        for indicator in self.indicators:
            if indicator.i_cal != 'sd' and indicator.chi != None:
                print(f'{indicator.name} analysis starts (chi2)')
                sheet_name = indicator.indicator_name
                var_name = indicator.description
                df = indicator.df
                dfs = {}
                dis_cols = list(indicator.breakdown.keys())
                book = load_workbook(file_path)
            
                for col, i in zip(dis_cols, range(len(dis_cols))):
                    f_df = chi2(df, indicator.var, col, alpha = 0.05)
                    if f_df is not None and not f_df.empty:
                        dfs[f'final_df{i}'] = f_df
                    
                if len(dfs) > 0:
                    final_df = pd.concat(dfs, axis=0)
                else:
                    final_df = pd.DataFrame()    
            
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                    final_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                
                wb = load_workbook(file_path)
                ws = wb[sheet_name]
                ws.insert_rows(1)
                ws['A1'] = var_name
                ws['A1'].font = Font(bold=True)
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
    
                wb.save(file_path)
        
    def calculation(self, indicator, method):
        """
        - To create a new column based on the calculation conditions of the indicators 
          to serve as the basis for data visualization and analysis
        - Perform the calculations using the score_map from the Indicator class
        indicator: indicator class, Indicator from indicator class (bodhi_indicator)
        method: str, How to calculate this indicator?
                0. None: Descriptive Statistics
                1. 'score': Evaluate whether a data point is valid based on the value of the related columns
                2. 'divide': Categorise (or group) data points based on the value of the related columns
                3. 'score_average': Assess the validity of each data point based on the average value of the related columns
                4. 'score_sum': Determine the validity of each data point based on the total sum of the related columns value
                5. 'score_select': Calculate how many of the related columns were answered correctly by respondents
                6. 'score_select_allyes': Check if all related columns were answered with "Yes"
                7. 'score_select_allno': Check if all related columns were answered with "No"
                8. 'score_select_anyyes': Check if at least one of the related columns was answered with "Yes"
                9. 'score_select_anyno': Check if at least one of the related columns was answered with "No"
                10. 'score_select_manual': Manual Code for multiple selecting
        """
        df = indicator.df
        variable = indicator.name
        
        if method == "score":
            if indicator.score_map != None:
                df['score'] = df[indicator.var].map(indicator.score_map)
                df[variable] = df['score'].apply(lambda x: 'Pass' if x >= indicator.valid_point else 'Not Pass')
                df.drop(columns=['score'], inplace = True)
            else: df[variable] = df[indicator.var].apply(lambda x: 'Pass' if x >= indicator.valid_point else 'Not Pass')
            
        elif method == "score_average":
            if indicator.score_map != None:
                def scoring(row):
                    score = 0
                    columns = indicator.var
                    for col in columns:
                        value = row[col]
                        score += indicator.score_map.get(value, 0)
                    return score / len(columns)
                df['score'] = df.apply(scoring, axis=1)
                df[variable] = df['score'].apply(lambda x: 'Pass' if x >= indicator.valid_point else 'Not Pass')
                df.drop(columns=['score'], inplace = True)
            else: print("Please assign the score map for calculation")

        elif method == "score_sum":
            if indicator.score_map != None:
                def scoring(row):
                    score = 0
                    columns = indicator.var
                    for col in columns:
                        score += indicator.score_map.get(row[col], 0)
                    return score
                df['score'] = df.apply(scoring, axis=1)
                df[variable] = df['score'].apply(lambda x: 'Pass' if x >= indicator.valid_point else 'Not Pass')
                df.drop(columns=['score'], inplace = True)
            else: print("Please assign the score map for calculation")  
        
        indicator.var = variable
        indicator.df = df
        indicator.var_type = 'single'
            
        if indicator.var_change is not None:
            df[variable] = df[variable].replace(indicator.var_change)
            
    def indicator_analysis(self):
        """
        - To run the calculation function for all indicators
        """         
        for indicator in self.indicators:
            if indicator.condition is None:
                pass
            else:
                df_copy = indicator.df.copy()
                df_copy = df_copy[indicator.condition]
                indicator.df = df_copy

            if indicator.i_cal != None:
                self.calculation(indicator, indicator.i_cal)
        return print("All indicators have been calculated")
    
    def breakdown_count_bar(self, indicator, df, colname, file_path, figsize=(12, 8), rotation=0, fontsize=12):
        """
        - To generate bar plots through the breakdown data (Count only)
        indicator: indicator class, Indicator from indicator class (bodhi_indicator)
        df: Dataframe, Disaggregated dataframe
        colname: str, Reference column for the breakdown
        file_path: str, Directory where plots will be saved
        figsize: tuple, Size of plots
        rotation: int, Rotation angle for the x-axis ticks
        fontsize: int, Font size for plots
        """         
        breakdown = indicator.breakdown[colname]
        palette = [bodhi_complement, bodhi_blue, bodhi_tertiary, bodhi_primary_1, bodhi_grey, bodhi_secondary]
        if indicator.var_order != None:
            df = df.loc[indicator.var_order]
        ax = df.plot(kind='bar', stacked=False, width=0.6, figsize=figsize, color=palette)
        title = f'{indicator.description}\nby {breakdown}'
        output_file = f'{file_path}_{indicator.indicator_name}_{breakdown}_count.png'
        
        ax.set_ylabel('Count')
        ax.set_title(title)
        column_totals = df.sum(axis=0)
        
        bar_width = ax.patches[0].get_width()
        fontsize_auto = max(fontsize * (bar_width / 0.85), 8)
        
        for i, column in enumerate(df.columns):
            for j in range(len(df)):
                bar_index = i * len(df) + j
                bar = ax.patches[bar_index]
                bar_x = bar.get_x() + bar.get_width() / 2
                bar_height = bar.get_height()
                value = df[column].iloc[j]
                percentage = (value / column_totals[column]) * 100
                text = f'{value}\n({percentage:.1f}%)'
                ax.text(bar_x, bar_height, text, ha='center', va='bottom', fontsize=fontsize_auto)

        if indicator.i_type == 'Count' :          
            if indicator.target is not None:
               ax.axhline(y=indicator.target, color='red', linestyle='--', linewidth=0.5, label='Target')
            if indicator.baseline is not None:
                ax.axhline(y=indicator.baseline, color='blue', linestyle='--', linewidth=0.5, label='Baseline')
            if indicator.midline is not None:
                ax.axhline(y=indicator.midline, color='green', linestyle='--', linewidth=0.5,  label='Midline')

        def replace_spaces(text):
            delimiters = [' ', '/']
            if len(text) >= 13:
                spaces = [i for i, char in enumerate(text) if char in delimiters]
                if len(spaces) >= 3:
                    text = text[:spaces[0]] + '\n' + text[spaces[0] + 1:spaces[2]] + '\n' + text[spaces[2] + 1:]
                elif len(spaces) >= 1:
                    text = text[:spaces[0]] + '\n' + text[spaces[0]+1:]
            return text
        
        plt.title(title, fontsize=fontsize + 4)
        plt.xlabel(" ", fontsize=fontsize)
        plt.ylabel("Count", fontsize = fontsize)
        try:
            df.index = df.index.map(lambda x: x if isinstance(x, str) else ''.join(x))
        except TypeError: pass
        labels = [''.join(label) if isinstance(label, tuple) else label for label in df.index]
        labels = [replace_spaces(label) for label in labels]
        ax.set_xticklabels(labels, rotation=rotation, fontsize=fontsize)
        plt.legend(title=f'{breakdown} and Target', fontsize=fontsize-1)
        max_height = df.max().max()
        plt.ylim(0, max_height * 1.1)
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        plt.savefig(output_file, bbox_inches='tight', dpi=800)

    def breakdown_percentage_bar(self, indicator, df, colname, file_path, figsize=(12, 8), rotation=0, fontsize=12):
        """
        - To generate bar plots through the breakdown data (Percentage only)
        indicator: indicator class, Indicator from indicator class (bodhi_indicator)
        df: Dataframe, Disaggregated dataframe
        colname: str, Reference column for the breakdown
        file_path: str, Directory where plots will be saved
        figsize: tuple, Size of plots
        rotation: int, Rotation angle for the x-axis ticks
        fontsize: int, Font size for plots
        """      
        breakdown = indicator.breakdown[colname]
        palette = [bodhi_complement, bodhi_blue, bodhi_tertiary, bodhi_primary_1, bodhi_grey, bodhi_secondary]
        if indicator.var_order != None:
            df = df.loc[indicator.var_order]
        ax = df.plot(kind='bar', stacked=False, width=0.6, figsize=figsize, color=palette)
        title = f'{indicator.description}\nby {breakdown}'
        output_file = f'{file_path}_{indicator.indicator_name}_{breakdown}_percent.png'
        
        ax.set_ylabel('Percentage')
        ax.set_title(title)
        bar_width = ax.patches[0].get_width()
        fontsize_auto = max(fontsize * (bar_width / 0.85), 8)
        for i in ax.containers:
            ax.bar_label(i, labels=[f'{p:.0f}%' for p in df[i.get_label()]], label_type='edge', fontsize=fontsize_auto)
            
        if indicator.i_type == 'Percentage' :          
            if indicator.target is not None:
               ax.axhline(y=indicator.target, color='red', linestyle='--', linewidth=0.5, label='Target')
            if indicator.baseline is not None:
                ax.axhline(y=indicator.baseline, color='blue', linestyle='--', linewidth=0.5, label='Baseline')
            if indicator.midline is not None:
                ax.axhline(y=indicator.midline, color='green', linestyle='--', linewidth=0.5,  label='Midline')
                
        def replace_spaces(text):
            delimiters = [' ', '/']
            if len(text) >= 13:
                spaces = [i for i, char in enumerate(text) if char in delimiters]
                if len(spaces) >= 3:
                    text = text[:spaces[0]] + '\n' + text[spaces[0] + 1:spaces[2]] + '\n' + text[spaces[2] + 1:]
                elif len(spaces) >= 1:
                    text = text[:spaces[0]] + '\n' + text[spaces[0]+1:]
            return text
    
        plt.title(title, fontsize=fontsize + 4)
        plt.xlabel(" ", fontsize=fontsize)
        plt.ylabel("Percentage", fontsize = fontsize)
        plt.ylim(0, 105)
        plt.yticks([0, 20, 40, 60, 80, 100])
        df.index = df.index.map(lambda x: x if isinstance(x, str) else ''.join(x))
        labels = [''.join(label) if isinstance(label, tuple) else label for label in df.index]
        labels = [replace_spaces(label) for label in labels]
        ax.set_xticklabels(labels, rotation=rotation, fontsize=fontsize)
        plt.xticks(rotation=rotation, fontsize=fontsize)
        plt.legend(title=f'{breakdown} and Target', fontsize=fontsize-1)
        plt.savefig(output_file, bbox_inches='tight', dpi=800)

    def plot_bar(self, indicator, df_, file_path, figsize=(12, 8), rotation=0, fontsize=12):
        """
        - To generate bar plot for overall information
        indicator: indicator class, Indicator from indicator class (bodhi_indicator)
        df_: Dataframe, Dataframe
        file_path: str, Directory where a bar plot will be saved
        figsize: tuple, Size of plots
        rotation: int, Rotation angle for the x-axis ticks
        fontsize: int, Font size for plots
        """      
        title = indicator.description
        output_file = f'{file_path}_{indicator.indicator_name}.png'
        palette = [bodhi_complement, bodhi_blue, bodhi_tertiary, bodhi_primary_1, bodhi_grey, bodhi_secondary]
        if indicator.var_order != None:
            df_ = df_.loc[indicator.var_order]        
        fig, ax = plt.subplots(figsize=figsize)
        if indicator.i_type == 'Count':
            df2 = df_['Count']
            df2.plot(kind='bar', color=palette, figsize=figsize, ax = ax)
            bars = ax.patches
            total = df_['Count'].values.sum()
            for bar in bars:
                height = bar.get_height()
                percentage = (height / total) * 100
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), 
                     f'{height:.0f} ({percentage:.1f}%)',
                     ha='center', va='bottom', fontsize=fontsize+2)
            max_height = df_.max().max()
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            plt.ylim(0, max_height * 1.1)
            
        elif indicator.i_type == 'Percentage':
            df_['Percentage'].plot(kind='bar', color=palette, figsize=figsize, ax = ax)
            bars = ax.patches
            for bar, (idx, row) in zip(bars, df_.iterrows()):
                percentage = row['Percentage']
                count = row['Count']
                label = f'{percentage:.1f}% ({int(count)})'
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), label,ha='center', va='bottom', fontsize=fontsize + 2)
            plt.ylim(0, 105)
            plt.yticks([0, 20, 40, 60, 80, 100])

        if indicator.target is not None:
            ax.axhline(y=indicator.target, color='red', linestyle='--', linewidth=0.5, label='Target')
            
        if indicator.baseline is not None:
            ax.axhline(y=indicator.baseline, color='blue', linestyle='--', linewidth=0.5, label='Baseline')
            
        if indicator.midline is not None:
            ax.axhline(y=indicator.midline, color='green', linestyle='--', linewidth=0.5,  label='Midline')
            
        df_.index = df_.index.map(lambda x: x if isinstance(x, str) else ''.join(x))
        labels = [''.join(label) if isinstance(label, tuple) else label for label in df_.index]

        def replace_spaces(text):
            delimiters = [' ', '/']
            if len(text) >= 13:
                spaces = [i for i, char in enumerate(text) if char in delimiters]
                if len(spaces) >= 3:
                    text = text[:spaces[0]] + '\n' + text[spaces[0] + 1:spaces[2]] + '\n' + text[spaces[2] + 1:]
                elif len(spaces) >= 1:
                    text = text[:spaces[0]] + '\n' + text[spaces[0]+1:]
            return text

        labels = [replace_spaces(label) for label in labels]
        plt.title(title, fontsize=fontsize + 4)
        plt.xlabel(" ", fontsize=fontsize)
        plt.ylabel(indicator.i_type, fontsize = fontsize)
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=rotation, fontsize=fontsize)
        bar_handles = [Patch(color=palette[i], label=label) for i, label in enumerate(labels)]
        line_handles, _ = ax.get_legend_handles_labels()
        handles = bar_handles + line_handles[:-1]
        ax.legend(handles=handles, title="Category", loc='best')
        plt.savefig(output_file, bbox_inches='tight', dpi=800)      

    def kap_tables(self, indicator, var, sheet_name, var_name, file_path, folder):
        """
        - To generate tables including both general and breakdown data and related plots
        indicator: indicator class, Indicator from indicator class (bodhi_indicator)
        var: list, Variables related to the indicator or question
        sheet_name: str, Excel sheet name for KAP analysis outputs
        var_name: str, Name of the question
        file_path: str, Directory where tables files will be saved
        folder: str, Folder where plots will be saved
        """
        df = indicator.df
        if indicator.breakdown != None:
            dis_cols = list(indicator.breakdown.keys())
        else: dis_cols = None
        dfs = {}
        book = load_workbook(file_path)  

        try:
            if indicator.var_order is not None:
                for var_ in var:
                    df[var_] = df[var_].astype('category')
                    df[var_] = df[var_].cat.set_categories(indicator.var_order, ordered=True)
        except (KeyError, ValueError) as e:
            print("")
            
        if indicator.var_type == 'single':
            overall_df = self.count(df, var, index_name=indicator.indicator_name)
            if indicator.kap_label == None:
                self.plot_bar(indicator, overall_df, folder)
            elif len(indicator.kap_label) < 7:
                self.plot_bar(indicator, overall_df, folder)
                
        elif indicator.var_type == 'multi':
            if indicator.var_change != None:
                change = list(indicator.var_change.values())
                overall_df = self.multi_table(indicator.df, indicator.var, 
                          categories = indicator.var_order, column_labels = indicator.kap_label, index_name = var_name, change=change)
            else:
                overall_df = self.multi_table(indicator.df, indicator.var, 
                      categories = indicator.var_order, column_labels = indicator.kap_label, index_name = var_name)
                
        if dis_cols != None:
            melted = df.melt(id_vars=dis_cols, value_vars=var, var_name=' ', value_name='category_value')
            for col, i in zip(dis_cols, range(len(dis_cols))):
                count_df = melted.groupby(['category_value', col]).size().unstack(fill_value=0)
                if indicator.var_type != 'multi':
                    self.breakdown_count_bar(indicator, count_df, col, folder)
                percent_df = round(count_df.div(count_df.sum(axis=0), axis=1) * 100, 2)
                if indicator.var_type != 'multi':
                    self.breakdown_percentage_bar(indicator, percent_df, col, folder)
                f_df = pd.concat([count_df, percent_df.add_suffix('(%)')], axis=1)
                breakdown = indicator.breakdown[col]
                dfs[f'{breakdown}'] = f_df.transpose()
                      
            final_df = pd.concat(dfs, axis=0)
            if indicator.var_change != None:
                final_df.rename(columns=indicator.var_change, inplace=True)
            
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            if dis_cols != None:
                final_df.to_excel(writer, sheet_name=sheet_name, merge_cells=False, index=True, header=True)
                startrow = final_df.shape[0] + 2
                overall_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=True, header=True)
            else: overall_df.to_excel(writer, sheet_name=sheet_name, index=True, header=True)
                
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        ws.insert_rows(1)
        ws['B1'] = indicator.description
        ws['B1'].font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
        wb.save(file_path)

    def kap(self, file_path, folder):
        """
        - Function to run the kap_tables function for each indicator or question (KAP)
        file_path: str, Directory where tables files will be saved
        folder: str, Folder where plots will be saved
        """
        for indicator in self.indicators:
            print(f'{indicator.name} analysis starts')
            if indicator.var_type == 'single' and indicator.chi == None:
               sheet_name = f"{indicator.indicator_name}"
               var_name = f"{indicator.number}" 
               self.kap_tables(indicator, indicator.var, sheet_name, var_name, file_path, folder)
            elif indicator.var_type == 'multi' and indicator.chi == None:
                names = range(len(indicator.var))
                for var, i in zip(indicator.var, names):
                    sheet_name = f"{indicator.indicator_name}-{i}"
                    var_name = f"{indicator.number}-{i}"
                    self.kap_tables(indicator, var, sheet_name, var_name, file_path, folder)
                    